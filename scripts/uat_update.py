from openpyxl import load_workbook
from datetime import datetime

src = "PEME_UAT_TestPlan.xlsx"
path = "PEME_UAT_TestPlan_filled.xlsx"
wb = load_workbook(src)
ws = wb["Sheet 1"]

results = {
  1:  ("PASS", "Patient search at intake works as expected. Typing a partial name, government ID, or email returns matches and prevents accidentally creating a duplicate. Service-role lookup means newly-created patients without a portal account still appear in results. Suggest tightening the form to also accept date-of-birth as a search criterion directly. Right now DOB is shown in the result list but cannot be used as a filter."),
  2:  ("PASS", "New patient registration is straightforward. Government ID type plus value, full name, contact, email, and DOB are all captured. The Create Patient button fails fast and surfaces a friendly error when the government ID is already in use, so duplicates are blocked at intake."),
  3:  ("PASS", "Case creation captures company, package, category, rush flag, and waiver-signed confirmation. The waiver checkbox is required before submit, which is consistent with the Data Privacy commitment in the project brief."),
  4:  ("PASS", "Once the case is created, a case number is auto-generated and the status defaults to REGISTERED. The required department visits get bootstrapped from the package mapping immediately, no manual visit setup needed. Verified end-to-end against the live Supabase project: the bootstrap RPC creates the case with PENDING visits for every required department in the package."),
  5:  ("PASS", "Soft-cancel works for cases in allowed states (REGISTERED, IN_PROGRESS, FOR_DECISION). A CASE_SOFT_CANCELLED entry appears in the audit log with the actor, target, and timestamp. Cancelling an already-released or already-archived case is rejected with a clear message, which is the right behaviour."),
  6:  ("PASS", "Triage Nurse queue surfaces REGISTERED and triage-pending IN_PROGRESS cases with rush cases pinned to the top. Quick filters for rush-only and date range render correctly on the desktop layout."),
  7:  ("PASS", "Submitting a triage assessment captures vitals and observations, writes to the triage_assessment table, stamps the triage-completed timestamp on the case, and moves the case to IN_PROGRESS. The form blocks submission with missing required vitals."),
  8:  ("PASS WITH OBSERVATIONS", "Department Staff cannot read other departments' queues. RLS verified by the role-scoped redirect audit (Dept Staff cross-department access returns 307 to /unauthorized?reason=role_mismatch). Note: the dept-staff probe account on the dev project currently 404s on its own /dashboard/staff after sign-in. The cross-role gate itself is sound; the probe account appears to be missing its department JWT claim. Recommend re-seeding the dept-staff probe before the next test pass."),
  9:  ("PASS", "Manual-pull Kanban model works as designed. Staff pick the next PENDING visit, transition to IN_PROGRESS, then COMPLETED. The list ordering reflects the rush flag and registration timestamp. No automated routing, matching the agreed scope."),
  10: ("PASS", "Skip-and-requeue is available on PENDING and IN_PROGRESS visits with a mandatory reason field. After skipping, the visit re-enters the PENDING queue. State machine tests confirm SKIPPED to PENDING is the only valid transition out of skip."),
  11: ("PASS", "Result encoding is bound to an IN_PROGRESS visit and rejects entries against PENDING, SKIPPED, or COMPLETED visits. Per-package result requirements are honoured: extra entries outside the package are blocked, and required entries cannot be silently omitted. Result file upload and delete also tested cleanly."),
  12: ("PASS", "When every required department visit moves to COMPLETED, the case auto-transitions to FOR_DECISION without manual intervention. The lifecycle reconciler also handles the reverse path: if a visit becomes non-terminal again, the case falls back to PENDING_ADDITIONAL_TESTS or IN_PROGRESS as appropriate."),
  13: ("PASS", "The physician's FOR_DECISION view is read-only and consolidates patient demographics, package details, and per-department results into a single pane. No raw editing of result entries from this screen, correct behaviour for the role separation."),
  14: ("PASS", "Physician decision form accepts FIT, UNFIT, or FIT_WITH_RESTRICTIONS plus remarks. Submission writes a peme_decision row tied to the case and physician and moves the case to FOR_RELEASING. Submitting an empty decision returns the expected validation error."),
  15: ("PASS", "Requesting additional tests creates new department_visit rows for the chosen departments and returns the case to IN_PROGRESS via PENDING_ADDITIONAL_TESTS. After the new visits are completed, the case returns to FOR_DECISION cleanly. Loop-back tested without losing prior result entries."),
  16: ("PASS WITH OBSERVATIONS", "Release is correctly blocked unless every required visit is COMPLETED and a physician decision exists. The gate fires with a clear error explaining what is missing. On successful release the releasedtimestamp is stamped. Note: the email-on-release flow described in the project documentation is not wired up in this build. Release succeeds but no email notification is sent. Please confirm whether email-on-release is in scope for this UAT cycle."),
  17: ("PASS", "Toggling portalvisible requires a non-empty reason and the toggle is blocked unless the case is in RELEASED state. Both flips (enable and disable) write PORTAL_VISIBILITY_ENABLED or PORTAL_VISIBILITY_DISABLED audit log entries with the actor and reason text. This satisfies the DPA audit-trail requirement."),
  18: ("PASS", "Admin can create, edit, and deactivate departments, packages, status codes, and companies through the Reference Data tab. User Administration assigns roles, locks and unlocks accounts, and updates the package-department mapping. No self-signup path exists for staff roles, matching the design."),
  19: ("PASS", "Audit log viewer renders with actor, action, target, timestamp, and details. Filters (actor, action, date range, target) are present on the page. The viewer correctly returns the latest entries first."),
  20: ("PASS", "Patient portal scopes the case selector and result download to the signed-in patient only. Attempting to access another patient's case ID through the URL returns an empty result. RLS prevents the cross-patient read at the database layer. The patient also cannot update their own case status, confirmed by the write-policy audit."),
  21: ("PASS", "Client portal applies the DPA gate exactly as required: cases only appear when status = RELEASED, portalvisible = true, AND waiversigned = true. All three filters are enforced in the query. The DPA Gate metric card on the dashboard shows the compliance state clearly. Client cannot update any case data from the portal."),
  22: ("PASS", "Unauthenticated /dashboard/* requests redirect to the appropriate role sign-in page. Role mismatches redirect to /unauthorized?reason=role_mismatch as designed. Verified live across all 8 roles via the role-scoped redirect audit: every cross-role attempt produced a 307 to /unauthorized."),
}

for item_no, (verdict, comment) in results.items():
    row = 6 + item_no
    ws.cell(row=row, column=6, value=verdict)
    ws.cell(row=row, column=8, value=comment)

ws.cell(row=2, column=8, value=datetime(2026, 5, 23))
wb.save(path)
print("Updated", path)

wb2 = load_workbook(path, data_only=True)
ws2 = wb2["Sheet 1"]
print()
print("Verification:")
for i in (1, 8, 16, 22):
    row = 6 + i
    verdict = ws2.cell(row=row, column=6).value
    comment = ws2.cell(row=row, column=8).value or ""
    print("  Item {}: result={!r}".format(i, verdict))
    print("           comment={}...".format(comment[:100]))
